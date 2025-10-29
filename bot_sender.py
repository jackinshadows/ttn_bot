import asyncio
import glob
import os
import sys
import datetime
from aiogram import Bot
from aiogram.types import FSInputFile
from openpyxl import load_workbook

# ---------- Налаштування ----------
TOKEN = "8404246224:AAGTrWXTqoFgt4Xe2OiPWMNX91oOdBsYcEc"            # встав свій токен
GROUP_ID = -1003118858109          # встав ID групи
FOLDER_PATH = r"C:\TTN"            # де лежать docx файли
EXCEL_PATH = r"C:\TTN\Швидкі ттн v1.2.xlsx"  # файл з F2 = ім'я менеджера
LOG_PATH = r"C:\TTN\ttn_log.txt"   # куди писати лог
# ------------------------------------

def write_log(message: str):
    """Запис повідомлення у лог з міткою часу."""
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        f.write(f"[{timestamp}] {message}\n")
    print(message)

async def send_latest_ttn():
    write_log("=== Запуск скрипта ===")

    if TOKEN == "ВАШ_BOT_TOKEN":
        write_log("❌ Не вказано BOT TOKEN.")
        return
    if GROUP_ID == -1001234567890:
        write_log("❌ Не вказано GROUP_ID.")
        return

    # 1) Читаємо ім'я менеджера з Excel
    manager_name = "Невідомий менеджер"
    try:
        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH, data_only=True)
            sheet_name = "Технічний лист" if "Технічний лист" in wb.sheetnames else wb.sheetnames[0]
            sheet = wb[sheet_name]
            val = sheet["F2"].value
            if val and str(val).strip():
                manager_name = str(val).strip()
            write_log(f"Ім'я менеджера: {manager_name}")
            wb.close()
        else:
            write_log(f"⚠️ Excel файл не знайдено: {EXCEL_PATH}")
    except Exception as e:
        write_log(f"❌ Помилка при читанні Excel: {e}")

    # 2) Шукаємо .docx файли (крім тимчасових ~$)
    pattern = os.path.join(FOLDER_PATH, "*.docx")
    all_files = glob.glob(pattern)
    write_log(f"Знайдено файлів у папці: {len(all_files)}")
    for f in all_files:
        write_log(f"  -> {os.path.basename(f)}")

    files = [f for f in all_files if not os.path.basename(f).startswith("~$")]
    write_log(f"Після фільтру (~$): {len(files)}")

    if not files:
        write_log("❌ Немає файлів для відправки.")
        return

    latest_file = max(files, key=lambda f: os.path.getmtime(f))
    write_log(f"Обрано файл: {latest_file}")

    # 3) Відправка у Telegram
    bot = Bot(token=TOKEN)
    try:
        doc = FSInputFile(latest_file)
        file_name = os.path.basename(latest_file)
        caption = f"📄 ТТН від: {manager_name}\nФайл: {file_name}"
        write_log(f"Відправляю файл '{file_name}' у групу...")
        await bot.send_document(chat_id=GROUP_ID, document=doc, caption=caption)
        write_log("✅ Успішно відправлено в Telegram.")
    except Exception as e:
        write_log(f"❌ Помилка при відправці в Telegram: {e}")
    finally:
        try:
            await bot.session.close()
        except Exception:
            pass

    write_log("=== Завершення скрипта ===\n")

if __name__ == "__main__":
    if len(sys.argv) >= 2:
        FOLDER_PATH = sys.argv[1]
    if len(sys.argv) >= 3:
        EXCEL_PATH = sys.argv[2]
    asyncio.run(send_latest_ttn())